import csv
import io
import json
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _build_performance_matrix(daily_performance: list[dict]) -> tuple:
    """Build date-controller performance matrix from raw daily data."""
    date_map = {}
    controller_map = defaultdict(dict)

    for row in daily_performance:
        date_obj = row["date_lot"]
        date_str = date_obj.strftime("%d/%m/%Y") if date_obj else "N/A"
        controller = row["controleur"]
        lots = int(row["total_lots"] or 0)
        actes = int(row["total_actes"] or 0)

        if date_str not in date_map:
            date_map[date_str] = date_obj

        controller_map[controller][date_str] = {"lots": lots, "actes": actes}

    sorted_dates = sorted(date_map.keys(), key=lambda d: date_map[d] or datetime.min)
    sorted_controllers = sorted(controller_map.keys())

    return sorted_dates, sorted_controllers, controller_map


def generate_csv(
    general_stats: dict,
    controleur_stats: list[dict],
    daily_performance: list[dict],
) -> str:
    """Generate a CSV report string."""
    sorted_dates, sorted_controllers, controller_map = _build_performance_matrix(
        daily_performance
    )

    lines = []

    # Header row
    header = ["Chef d'équipe"]
    for date in sorted_dates:
        header.append(f"{date} - Lots")
        header.append(f"{date} - Actes")
    lines.append(",".join(header))

    # Data rows + accumulate totals
    daily_lots_total = defaultdict(int)
    daily_actes_total = defaultdict(int)

    for controller in sorted_controllers:
        row = [controller]
        for date in sorted_dates:
            data = controller_map[controller].get(date, {"lots": 0, "actes": 0})
            row.append(str(data["lots"]) if data["lots"] else "")
            row.append(str(data["actes"]) if data["actes"] else "")
            daily_lots_total[date] += data["lots"]
            daily_actes_total[date] += data["actes"]
        lines.append(",".join(row))

    # Total row
    total_row = ["Total général"]
    for date in sorted_dates:
        total_row.append(str(daily_lots_total[date]))
        total_row.append(str(daily_actes_total[date]))
    lines.append(",".join(total_row))
    lines.append("")
    lines.append("")

    # Error statistics table
    lines.append(
        "Chef d'équipe,Nbr d'image Controlee,Nbr d'erreur détecté,Taux d'erreur"
    )
    for c in controleur_stats:
        lines.append(
            f"{c['controleur']},{c['total_actes_controlees']},{c['total_erreurs']},{c['taux_erreur']}%"
        )

    return "\ufeff" + "\n".join(lines)


def generate_excel(
    general_stats: dict,
    controleur_stats: list[dict],
    daily_performance: list[dict],
) -> bytes:
    """Generate a styled Excel report and return as bytes."""
    sorted_dates, sorted_controllers, controller_map = _build_performance_matrix(
        daily_performance
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Performance"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(vertical="center", horizontal="center")
    left_align = Alignment(vertical="center", horizontal="left")
    thin_border = Border(
        top=Side(style="thin"),
        left=Side(style="thin"),
        bottom=Side(style="thin"),
        right=Side(style="thin"),
    )
    medium_border = Border(
        top=Side(style="medium"),
        left=Side(style="thin"),
        bottom=Side(style="medium"),
        right=Side(style="thin"),
    )

    def style_header_cell(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # === TABLE 1: Daily Performance Matrix ===
    # Row 1: Date headers (merged)
    ws.cell(row=1, column=1, value="Chef d'equipe")
    style_header_cell(ws.cell(row=1, column=1))

    for idx, date in enumerate(sorted_dates):
        start_col = 2 + idx * 2
        ws.cell(row=1, column=start_col, value=date)
        ws.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1
        )
        style_header_cell(ws.cell(row=1, column=start_col))
        style_header_cell(ws.cell(row=1, column=start_col + 1))

    total_col_start = 2 + len(sorted_dates) * 2
    ws.cell(row=1, column=total_col_start, value="Total")
    ws.merge_cells(
        start_row=1,
        start_column=total_col_start,
        end_row=1,
        end_column=total_col_start + 1,
    )
    style_header_cell(ws.cell(row=1, column=total_col_start))
    style_header_cell(ws.cell(row=1, column=total_col_start + 1))

    # Row 2: Sub-headers (Lots / Actes)
    ws.cell(row=2, column=1, value="")
    style_header_cell(ws.cell(row=2, column=1))

    for idx in range(len(sorted_dates)):
        start_col = 2 + idx * 2
        ws.cell(row=2, column=start_col, value="Lots")
        ws.cell(row=2, column=start_col + 1, value="Actes")
        style_header_cell(ws.cell(row=2, column=start_col))
        style_header_cell(ws.cell(row=2, column=start_col + 1))

    ws.cell(row=2, column=total_col_start, value="Actes")
    ws.cell(row=2, column=total_col_start + 1, value="Lots")
    style_header_cell(ws.cell(row=2, column=total_col_start))
    style_header_cell(ws.cell(row=2, column=total_col_start + 1))

    # Data rows
    daily_lots_total = defaultdict(int)
    daily_actes_total = defaultdict(int)
    current_row = 3

    for controller in sorted_controllers:
        ws.cell(row=current_row, column=1, value=controller)
        ws.cell(row=current_row, column=1).alignment = left_align
        ws.cell(row=current_row, column=1).border = thin_border

        ctrl_total_lots = 0
        ctrl_total_actes = 0

        for idx, date in enumerate(sorted_dates):
            start_col = 2 + idx * 2
            data = controller_map[controller].get(date, {"lots": 0, "actes": 0})

            cell_lots = ws.cell(
                row=current_row, column=start_col, value=data["lots"] or ""
            )
            cell_actes = ws.cell(
                row=current_row, column=start_col + 1, value=data["actes"] or ""
            )
            cell_lots.alignment = center_align
            cell_lots.border = thin_border
            cell_actes.alignment = center_align
            cell_actes.border = thin_border

            ctrl_total_lots += data["lots"]
            ctrl_total_actes += data["actes"]
            daily_lots_total[date] += data["lots"]
            daily_actes_total[date] += data["actes"]

        # Total column for this controller
        cell = ws.cell(row=current_row, column=total_col_start, value=ctrl_total_actes)
        cell.alignment = center_align
        cell.border = thin_border
        cell = ws.cell(
            row=current_row, column=total_col_start + 1, value=ctrl_total_lots
        )
        cell.alignment = center_align
        cell.border = thin_border

        current_row += 1

    # Total général row
    ws.cell(row=current_row, column=1, value="Total general")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=1).border = medium_border

    grand_total_lots = 0
    grand_total_actes = 0

    for idx, date in enumerate(sorted_dates):
        start_col = 2 + idx * 2
        lots = daily_lots_total[date]
        actes = daily_actes_total[date]

        cell = ws.cell(row=current_row, column=start_col, value=lots)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = medium_border

        cell = ws.cell(row=current_row, column=start_col + 1, value=actes)
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = medium_border

        grand_total_lots += lots
        grand_total_actes += actes

    cell = ws.cell(row=current_row, column=total_col_start, value=grand_total_actes)
    cell.font = Font(bold=True)
    cell.alignment = center_align
    cell.border = medium_border
    cell = ws.cell(
        row=current_row, column=total_col_start + 1, value=grand_total_lots
    )
    cell.font = Font(bold=True)
    cell.alignment = center_align
    cell.border = medium_border

    current_row += 3  # spacing

    # === TABLE 2: Quality Metrics ===
    quality_headers = [
        "Chef d'equipe",
        "Nbr d'image Controlee",
        "Nbr d erreur detecte",
        "Taux d'erreur",
    ]
    for col_idx, header in enumerate(quality_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        style_header_cell(cell)

    current_row += 1

    for c in controleur_stats:
        ws.cell(row=current_row, column=1, value=c["controleur"])
        ws.cell(row=current_row, column=1).alignment = left_align
        ws.cell(row=current_row, column=1).border = thin_border

        ws.cell(row=current_row, column=2, value=c["total_actes_controlees"])
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=2).border = thin_border

        ws.cell(row=current_row, column=3, value=c["total_erreurs"])
        ws.cell(row=current_row, column=3).alignment = center_align
        ws.cell(row=current_row, column=3).border = thin_border

        ws.cell(row=current_row, column=4, value=f"{c['taux_erreur']}%")
        ws.cell(row=current_row, column=4).alignment = center_align
        ws.cell(row=current_row, column=4).border = thin_border

        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    total_columns = 1 + len(sorted_dates) * 2 + 2
    for i in range(2, total_columns + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_json_report(
    general_stats: dict,
    controleur_stats: list[dict],
    daily_performance: list[dict],
) -> dict:
    """Generate a JSON report structure."""
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    return {
        "metadata": {
            "date_generation": timestamp,
            "type_rapport": "Rapport de Performance Quotidienne",
            "periode": {
                "debut": str(general_stats.get("date_premiere", "")),
                "fin": str(general_stats.get("date_derniere", "")),
            },
        },
        "statistiques_generales": {
            "total_lots": int(general_stats.get("total_lots", 0)),
            "total_actes_traites": int(general_stats.get("total_actes_traites") or 0),
            "total_actes_rejets": int(general_stats.get("total_actes_rejets") or 0),
        },
        "performance_quotidienne": [
            {k: str(v) if isinstance(v, (datetime,)) else v for k, v in row.items()}
            for row in daily_performance
        ],
        "metriques_qualite": controleur_stats,
    }
